#!/usr/bin/env python3
"""
Transformateur Excel - Supprime les lignes N/A et les macros.

Règles de filtrage :
  - 2024 : supprime les lignes où la colonne I (N° facture CR) OU J (N° facture PI) = "N/A"
  - 2025 : supprime les lignes où la colonne F (N° de facture) = "N/A"
  - 2026 : supprime les lignes où la colonne F (N° de facture) = "N/A"

Résultat : fichier .xlsx propre, sans macro, avec la même présentation.
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
import os
import threading
from pathlib import Path


# ─── Configuration par année ───────────────────────────────────────────────────

FEUILLE_PRINCIPALE = {
    2024: "Base 2024",
    2025: "Feuil1",
    2026: "Base 2026",
}

# Numéros de colonnes (1-based) à vérifier pour "N/A"
COLONNES_NA = {
    2024: [9, 10],  # I = N° facture CR, J = N° facture PI
    2025: [6],       # F = N° de facture
    2026: [6],       # F = N° de facture
}


# ─── Logique de traitement ─────────────────────────────────────────────────────

def detecter_annee(nom_fichier: str):
    for annee in [2024, 2025, 2026]:
        if str(annee) in nom_fichier:
            return annee
    return None


def est_ligne_na(ws, numero_ligne: int, annee: int) -> bool:
    colonnes = COLONNES_NA.get(annee, [6])
    for col in colonnes:
        if ws.cell(row=numero_ligne, column=col).value == "N/A":
            return True
    return False


def derniere_ligne_reelle(ws) -> int:
    """Retourne le numéro de la dernière ligne contenant des données."""
    for ligne in range(ws.max_row, 1, -1):
        if any(ws.cell(row=ligne, column=c).value is not None
               for c in range(1, ws.max_column + 1)):
            return ligne
    return 1


def traiter_fichier(chemin_entree: str, log=None):
    """
    Traite un fichier Excel :
      - supprime les lignes N/A selon les règles de l'année
      - renuméroter la colonne A (#)
      - sauvegarde en .xlsx sans macro

    Retourne (chemin_sortie, nb_lignes_supprimées).
    """
    def journaliser(msg: str):
        if log:
            log(msg)

    nom = os.path.basename(chemin_entree)
    annee = detecter_annee(nom)

    if annee is None:
        raise ValueError(
            f"Impossible de détecter l'année (2024/2025/2026) dans le nom : {nom}\n"
            "Assurez-vous que l'année est dans le nom du fichier."
        )

    journaliser(f"📂 {nom}  →  année {annee}")

    # Chargement sans VBA (supprime macros et boutons)
    wb = openpyxl.load_workbook(chemin_entree, keep_vba=False)

    # Feuille principale
    nom_feuille = FEUILLE_PRINCIPALE.get(annee, wb.sheetnames[0])
    if nom_feuille not in wb.sheetnames:
        nom_feuille = wb.sheetnames[0]
        journaliser(f"  ⚠️  Feuille '{FEUILLE_PRINCIPALE.get(annee)}' introuvable, "
                    f"utilisation de '{nom_feuille}'")
    ws = wb[nom_feuille]

    journaliser(f"  Feuille : « {nom_feuille} »")

    # Vraie dernière ligne (évite les 17 000 lignes vides du 2026)
    fin = derniere_ligne_reelle(ws)
    journaliser(f"  Lignes de données : {fin - 1}")

    # Collecte des lignes à supprimer
    lignes_na = [
        r for r in range(2, fin + 1)
        if est_ligne_na(ws, r, annee)
    ]
    journaliser(f"  Lignes N/A trouvées : {len(lignes_na)}")

    # Suppression en ordre inverse pour ne pas décaler les indices
    for r in reversed(lignes_na):
        ws.delete_rows(r)

    # Supprimer les lignes vides excédentaires (ex: 2026 avait 17 000 lignes vides)
    nouvelle_fin = derniere_ligne_reelle(ws)
    if ws.max_row > nouvelle_fin + 5:
        ws.delete_rows(nouvelle_fin + 1, ws.max_row - nouvelle_fin)

    # Renumérotation de la colonne A (le champ "#")
    compteur = 1
    for r in range(2, ws.max_row + 1):
        cel = ws.cell(row=r, column=1)
        if cel.value is not None:
            cel.value = compteur
            compteur += 1

    # Nom du fichier de sortie
    racine = Path(chemin_entree).stem
    # Enlever l'extension .xlsm éventuelle du stem
    if racine.endswith(".xlsm"):
        racine = racine[:-5]
    nom_sortie = racine + "_officiel.xlsx"
    dossier_sortie = os.path.dirname(chemin_entree)
    chemin_sortie = os.path.join(dossier_sortie, nom_sortie)

    wb.save(chemin_sortie)

    journaliser(f"  ✅ Sauvegardé : {nom_sortie}")
    journaliser(f"     {len(lignes_na)} ligne(s) supprimée(s) "
                f"| {compteur - 1} ligne(s) conservée(s)\n")

    return chemin_sortie, len(lignes_na)


# ─── Interface graphique ───────────────────────────────────────────────────────

VERT   = "#27ae60"
ROUGE  = "#e74c3c"
BLEU   = "#2980b9"
FOND   = "#f0f2f5"
SOMBRE = "#2c3e50"
GRIS   = "#95a5a6"


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Transformateur Excel – Suppression N/A")
        self.geometry("720x560")
        self.resizable(True, True)
        self.configure(bg=FOND)
        self.fichiers: list[str] = []
        self._construire_ui()

    # ── Construction de l'interface ──────────────────────────────────────────

    def _construire_ui(self):
        # Titre
        tk.Label(
            self, text="Transformateur Excel",
            font=("Arial", 20, "bold"), bg=FOND, fg=SOMBRE
        ).pack(pady=(22, 2))
        tk.Label(
            self,
            text="Supprime les lignes N/A · Retire les macros · Sauvegarde en .xlsx",
            font=("Arial", 11), bg=FOND, fg=GRIS
        ).pack(pady=(0, 18))

        # Zone de sélection de fichiers
        cadre_zone = tk.Frame(self, bg="#d5dde6", bd=0, relief="flat")
        cadre_zone.pack(fill="x", padx=40)

        cadre_interne = tk.Frame(cadre_zone, bg="#dfe6ed", bd=2, relief="groove",
                                  cursor="hand2")
        cadre_interne.pack(fill="x", padx=2, pady=2)

        lbl_zone = tk.Label(
            cadre_interne,
            text="Cliquer ici pour sélectionner des fichiers Excel\n(.xlsx  /  .xlsm)",
            font=("Arial", 12), bg="#dfe6ed", fg=SOMBRE, pady=18
        )
        lbl_zone.pack()

        for widget in (cadre_interne, lbl_zone):
            widget.bind("<Button-1>", self._parcourir_fichiers)

        # Liste des fichiers sélectionnés
        self._var_liste = tk.StringVar(value=[])
        self.listbox = tk.Listbox(
            self, listvariable=self._var_liste,
            height=4, font=("Arial", 10), bg="white",
            selectmode=tk.MULTIPLE, activestyle="none",
            relief="flat", bd=1, highlightthickness=1,
            highlightbackground="#bdc3c7"
        )
        self.listbox.pack(fill="x", padx=40, pady=(8, 0))

        # Boutons
        cadre_boutons = tk.Frame(self, bg=FOND)
        cadre_boutons.pack(pady=12)

        self.btn_traiter = tk.Button(
            cadre_boutons, text="▶  Traiter les fichiers",
            font=("Arial", 12, "bold"), bg=VERT, fg="white",
            padx=22, pady=9, relief="flat", cursor="hand2",
            activebackground="#219a52", activeforeground="white",
            command=self._lancer_traitement
        )
        self.btn_traiter.pack(side="left", padx=6)

        tk.Button(
            cadre_boutons, text="Effacer la liste",
            font=("Arial", 11), bg=ROUGE, fg="white",
            padx=16, pady=9, relief="flat", cursor="hand2",
            activebackground="#c0392b", activeforeground="white",
            command=self._effacer_liste
        ).pack(side="left", padx=6)

        tk.Button(
            cadre_boutons, text="Ouvrir le dossier",
            font=("Arial", 11), bg=BLEU, fg="white",
            padx=16, pady=9, relief="flat", cursor="hand2",
            activebackground="#2471a3", activeforeground="white",
            command=self._ouvrir_dossier
        ).pack(side="left", padx=6)

        # Journal
        cadre_journal = tk.Frame(self, bg=FOND)
        cadre_journal.pack(fill="both", expand=True, padx=40, pady=(0, 22))

        tk.Label(cadre_journal, text="Journal :", font=("Arial", 10, "bold"),
                 bg=FOND, fg=SOMBRE).pack(anchor="w")

        cadre_texte = tk.Frame(cadre_journal, bg=SOMBRE, bd=0)
        cadre_texte.pack(fill="both", expand=True)

        self.journal = tk.Text(
            cadre_texte, font=("Menlo", 10), bg="#1e2d3d", fg="#ecf0f1",
            state="disabled", relief="flat", padx=10, pady=8,
            insertbackground="white"
        )
        scrollbar = ttk.Scrollbar(cadre_texte, command=self.journal.yview)
        self.journal.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        self.journal.pack(side="left", fill="both", expand=True)

        # Tags de couleur pour le journal
        self.journal.tag_config("ok",    foreground="#2ecc71")
        self.journal.tag_config("erreur", foreground="#e74c3c")
        self.journal.tag_config("info",  foreground="#3498db")
        self.journal.tag_config("titre", foreground="#f39c12", font=("Menlo", 10, "bold"))

    # ── Actions ──────────────────────────────────────────────────────────────

    def _parcourir_fichiers(self, _event=None):
        fichiers = filedialog.askopenfilenames(
            title="Sélectionner des fichiers Excel",
            filetypes=[("Fichiers Excel", "*.xlsx *.xlsm"), ("Tous les fichiers", "*.*")]
        )
        for f in fichiers:
            if f not in self.fichiers:
                self.fichiers.append(f)
        self._actualiser_liste()

    def _actualiser_liste(self):
        noms = [os.path.basename(f) for f in self.fichiers]
        self._var_liste.set(noms)

    def _effacer_liste(self):
        self.fichiers.clear()
        self._actualiser_liste()
        self._journaliser("Liste effacée.", tag="info")

    def _ouvrir_dossier(self):
        if self.fichiers:
            dossier = os.path.dirname(self.fichiers[0])
        else:
            dossier = os.path.expanduser("~/Desktop")
        os.system(f'open "{dossier}"')

    def _journaliser(self, message: str, tag: str = ""):
        def _ecrire():
            self.journal.configure(state="normal")
            if tag:
                self.journal.insert("end", message + "\n", tag)
            else:
                self.journal.insert("end", message + "\n")
            self.journal.see("end")
            self.journal.configure(state="disabled")
        self.after(0, _ecrire)

    def _lancer_traitement(self):
        if not self.fichiers:
            messagebox.showwarning("Aucun fichier", "Veuillez sélectionner au moins un fichier Excel.")
            return

        self.btn_traiter.configure(state="disabled", text="⏳  Traitement en cours…")
        self._journaliser("═" * 55, tag="titre")
        self._journaliser("Début du traitement", tag="titre")
        self._journaliser("═" * 55, tag="titre")

        def traiter():
            total_supprimes = 0
            nb_ok = 0
            nb_erreurs = 0

            for chemin in self.fichiers:
                try:
                    _, supprimes = traiter_fichier(chemin, log=self._journaliser)
                    total_supprimes += supprimes
                    nb_ok += 1
                except Exception as exc:
                    self._journaliser(
                        f"❌ ERREUR – {os.path.basename(chemin)} : {exc}",
                        tag="erreur"
                    )
                    nb_erreurs += 1

            self._journaliser("═" * 55, tag="titre")
            self._journaliser(
                f"Terminé : {nb_ok} fichier(s) traité(s), "
                f"{total_supprimes} ligne(s) N/A supprimée(s)" +
                (f", {nb_erreurs} erreur(s)" if nb_erreurs else ""),
                tag="ok"
            )
            self._journaliser("═" * 55, tag="titre")

            self.after(0, lambda: self.btn_traiter.configure(
                state="normal", text="▶  Traiter les fichiers"
            ))

        threading.Thread(target=traiter, daemon=True).start()


# ─── Point d'entrée ────────────────────────────────────────────────────────────

if __name__ == "__main__":
    app = App()
    app.mainloop()
